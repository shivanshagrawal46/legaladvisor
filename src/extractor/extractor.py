"""
Unified facade for attachment text extraction.

Routes by extension:
  .pdf                 -> src.extractor.pdf  (text-layer + OCR fallback)
  .docx                -> src.extractor.docx
  .doc                 -> skipped (legacy binary; handled via word_com in v2)
  .xls                 -> xlrd (if installed)
  .msg                 -> extract_msg (if installed)
  .txt / .csv / .log   -> direct decode
  .png/.jpg/.jpeg/etc. -> src.extractor.image (PaddleOCR)
  .xlsx / .xls         -> openpyxl (.xlsx) / skipped (.xls)
  others (.zip, .exe)  -> skipped
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

from src.utils.logger import logger

# Image extensions supported by PaddleOCR
_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff", ".webp"}
_TEXT_EXTS = {".txt", ".csv", ".log", ".md", ".rtf"}


@dataclass
class PageResult:
    page_no: int
    text: str
    method: str  # "text_layer" | "ocr" | "raw" | "docx" | "xlsx"
    ocr_confidence: Optional[float] = None


@dataclass
class ExtractionResult:
    text: str
    method: str  # high-level: "pdf_text" | "pdf_ocr" | "pdf_mixed" | "docx" | "image_ocr" | "raw_text" | "xlsx" | "skipped"
    pages: List[PageResult] = field(default_factory=list)
    skipped_reason: Optional[str] = None
    char_count: int = 0
    avg_ocr_confidence: Optional[float] = None


def _filename_ext(filename: str) -> str:
    return Path(filename).suffix.lower()


def _decode_text(data: bytes) -> str:
    for enc in ("utf-8", "utf-16", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _extract_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return ""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return ""
    parts: List[str] = []
    for sheet in wb.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip() != ""]
            if cells:
                parts.append(" | ".join(cells))
    return "\n".join(parts).strip()


def _extract_xls(data: bytes) -> Optional[str]:
    """Legacy .xls via xlrd. Returns None if xlrd isn't installed (so the
    caller can mark a clear 'lib missing' reason rather than crash)."""
    try:
        import xlrd  # type: ignore
    except ImportError:
        return None
    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception:
        return ""
    parts: List[str] = []
    for sheet in book.sheets():
        parts.append(f"# Sheet: {sheet.name}")
        for r in range(sheet.nrows):
            cells = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            cells = [c for c in cells if c.strip()]
            if cells:
                parts.append(" | ".join(cells))
    return "\n".join(parts).strip()


def _extract_msg(data: bytes) -> Optional[str]:
    """Outlook .msg via extract_msg. Returns None if the lib isn't installed."""
    try:
        import extract_msg  # type: ignore
    except ImportError:
        return None
    import os
    import tempfile
    path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".msg", delete=False) as tf:
            tf.write(data)
            path = tf.name
        msg = extract_msg.Message(path)
        parts: List[str] = []
        for fld in ("subject", "sender", "to", "date"):
            v = getattr(msg, fld, None)
            if v:
                parts.append(f"{fld}: {v}")
        if getattr(msg, "body", None):
            parts.append(str(msg.body))
        return "\n".join(parts).strip()
    except Exception:
        return ""
    finally:
        if path:
            try:
                os.unlink(path)
            except OSError:
                pass


def extract_from_bytes(
    data: bytes,
    filename: str,
    *,
    ocr_lang: str = "en",
    ocr_min_chars: int = 80,
    ocr_dpi: int = 300,
    enable_ocr: bool = True,
    vision_enabled: bool = False,
    vision_model: str = "claude-sonnet-4-6",
    vision_min_pages: int = 3,
    vision_dpi: int = 180,
    vision_concurrency: int = 8,
) -> ExtractionResult:
    """Extract text from a binary blob. Routes by file extension."""
    ext = _filename_ext(filename)
    if not data:
        return ExtractionResult(text="", method="skipped", skipped_reason="empty")

    if ext == ".pdf":
        from src.extractor.pdf import extract_pdf

        pages = extract_pdf(
            data,
            ocr_lang=ocr_lang,
            ocr_min_chars=ocr_min_chars,
            ocr_dpi=ocr_dpi,
            enable_ocr=enable_ocr,
            vision_enabled=vision_enabled,
            vision_model=vision_model,
            vision_min_pages=vision_min_pages,
            vision_dpi=vision_dpi,
            vision_concurrency=vision_concurrency,
        )
        if not pages:
            return ExtractionResult(text="", method="skipped", skipped_reason="pdf_unreadable")

        n_text = sum(1 for p in pages if p.method == "text_layer" and p.text)
        n_ocr = sum(1 for p in pages if p.method in ("ocr", "claude_vision") and p.text)
        method = (
            "pdf_text" if n_ocr == 0
            else ("pdf_ocr" if n_text == 0 else "pdf_mixed")
        )
        confs = [p.ocr_confidence for p in pages if p.ocr_confidence is not None]
        avg_conf = sum(confs) / len(confs) if confs else None
        text = "\n\n".join(p.text for p in pages if p.text).strip()
        page_results = [
            PageResult(p.page_no, p.text, p.method, p.ocr_confidence) for p in pages
        ]
        return ExtractionResult(
            text=text,
            method=method,
            pages=page_results,
            char_count=len(text),
            avg_ocr_confidence=avg_conf,
        )

    if ext == ".docx":
        from src.extractor.docx import extract_docx

        text = extract_docx(data).strip()
        if not text:
            return ExtractionResult(text="", method="skipped", skipped_reason="docx_empty")
        return ExtractionResult(
            text=text,
            method="docx",
            pages=[PageResult(1, text, "docx")],
            char_count=len(text),
        )

    if ext in _IMAGE_EXTS:
        if not enable_ocr:
            return ExtractionResult(text="", method="skipped", skipped_reason="ocr_disabled")
        from src.extractor.image import extract_image

        text, conf = extract_image(data, lang=ocr_lang)
        text = text.strip()
        if not text:
            return ExtractionResult(text="", method="skipped", skipped_reason="image_no_text")
        return ExtractionResult(
            text=text,
            method="image_ocr",
            pages=[PageResult(1, text, "ocr", conf)],
            char_count=len(text),
            avg_ocr_confidence=conf,
        )

    if ext in _TEXT_EXTS:
        text = _decode_text(data).strip()
        if not text:
            return ExtractionResult(text="", method="skipped", skipped_reason="text_empty")
        return ExtractionResult(
            text=text,
            method="raw_text",
            pages=[PageResult(1, text, "raw")],
            char_count=len(text),
        )

    if ext == ".xlsx":
        text = _extract_xlsx(data).strip()
        if not text:
            return ExtractionResult(text="", method="skipped", skipped_reason="xlsx_empty")
        return ExtractionResult(
            text=text,
            method="xlsx",
            pages=[PageResult(1, text, "xlsx")],
            char_count=len(text),
        )

    if ext == ".xls":
        text = _extract_xls(data)
        if text is None:
            return ExtractionResult(text="", method="skipped", skipped_reason="xls_lib_missing")
        text = text.strip()
        if not text:
            return ExtractionResult(text="", method="skipped", skipped_reason="xls_empty")
        return ExtractionResult(text=text, method="xls",
                                pages=[PageResult(1, text, "xls")], char_count=len(text))

    if ext == ".msg":
        text = _extract_msg(data)
        if text is None:
            return ExtractionResult(text="", method="skipped", skipped_reason="msg_lib_missing")
        text = text.strip()
        if not text:
            return ExtractionResult(text="", method="skipped", skipped_reason="msg_empty")
        return ExtractionResult(text=text, method="msg",
                                pages=[PageResult(1, text, "msg")], char_count=len(text))

    if ext == ".doc":
        # Legacy binary Word. Clean extraction needs Word COM / LibreOffice;
        # the v2 pipeline already handles these via word_com. We do NOT do a
        # raw byte-scrape here (it injects garbage into evidence).
        return ExtractionResult(text="", method="skipped",
                                skipped_reason="legacy_doc_needs_word_com")

    return ExtractionResult(
        text="",
        method="skipped",
        skipped_reason=f"unsupported_ext:{ext or 'none'}",
    )


def extract_from_path(path: str | Path, **kwargs) -> ExtractionResult:
    p = Path(path)
    return extract_from_bytes(p.read_bytes(), p.name, **kwargs)
