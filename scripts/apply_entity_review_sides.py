"""Apply human-assigned entity sides from the reviewed workbook.

Reads `entity_review_for_jpsir_v4_updated.xlsx` (tab "Parties to label"),
column ASSIGNED_SIDE, and stamps each entity's side. Free-text notes in the
side cell (e.g. "david_network (David's Sister)") are parsed: the canonical
side is applied and the parenthetical is preserved as `side_note`.

  python -m scripts.apply_entity_review_sides            # DRY-RUN
  python -m scripts.apply_entity_review_sides --live     # apply
Idempotent.
"""
from __future__ import annotations

import re
import sys
from datetime import datetime, timezone

from openpyxl import load_workbook

from config.settings import Settings
from src.db.mongo import MongoClientWrapper
from src.graph.schema import SIDES
from src.utils.logger import logger

XLSX = "entity_review_for_jpsir_v4_updated.xlsx"
CANON = ["david_network", "our_side", "co_victim", "third_party", "unknown"]


def parse_side(raw: str):
    """Return (canonical_side, note) from a possibly-annotated cell."""
    if not raw:
        return None, None
    low = str(raw).strip().lower()
    for s in CANON:
        if low.startswith(s) or s in low:
            note = re.sub(r"^[^(]*\(?|\)?$", "", str(raw).strip())
            note = note.strip() if "(" in str(raw) else None
            # extract parenthetical specifically
            mo = re.search(r"\(([^)]*)\)", str(raw))
            return s, (mo.group(1).strip() if mo else None)
    return None, None


def main() -> int:
    live = "--live" in sys.argv
    now = datetime.now(timezone.utc)
    s = Settings.load()
    m = MongoClientWrapper(s.mongo_uri, s.mongo_db_name)
    ents = m.db["entities"]

    wb = load_workbook(XLSX, read_only=True)
    ws = wb["Parties to label"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    si = next(i for i, h in enumerate(hdr) if h and "ASSIGNED_SIDE" in str(h))
    ei = next(i for i, h in enumerate(hdr) if h and str(h).strip() == "entity_id")

    counts = {c: 0 for c in CANON}
    bad, missing, updates = [], [], []
    for r in rows[1:]:
        eid = r[ei]
        side, note = parse_side(r[si])
        if not eid or not r[si]:
            continue
        if side not in SIDES:
            bad.append((eid, r[si]))
            continue
        if not ents.find_one({"_id": eid}, {"_id": 1}):
            missing.append(eid)
            continue
        counts[side] += 1
        updates.append((eid, side, note))

    logger.info(f"=== apply entity sides ({'LIVE' if live else 'DRY-RUN'}) ===")
    logger.info(f"to apply: {sum(counts.values())} | by side: {counts}")
    if bad:
        logger.warning(f"unparseable side values ({len(bad)}): {bad[:8]}")
    if missing:
        logger.warning(f"entity_ids not found ({len(missing)}): {missing[:8]}")

    if live:
        for eid, side, note in updates:
            setd = {"side": side,
                    "is_david": side == "david_network",
                    "is_david_network": side == "david_network",
                    "is_ours": side == "our_side",
                    "side_source": "human_review_2026_06_17", "updated_at": now}
            if note:
                setd["side_note"] = note
            ents.update_one({"_id": eid}, {"$set": setd})
        logger.info(f"APPLIED {len(updates)} side assignments.")
        import collections
        dist = collections.Counter(e.get("side") for e in
                                   ents.find({"is_active": {"$ne": False}}, {"side": 1}))
        logger.info(f"active side distribution now: {dict(dist)}")
    else:
        logger.info("DRY-RUN — re-run with --live to apply.")
        for eid, side, note in updates[:5]:
            logger.info(f"   {side} <- {eid}" + (f"  note={note!r}" if note else ""))
    m.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
