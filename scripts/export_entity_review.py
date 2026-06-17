"""One-off export for human (CEO) review of entity data-hygiene items:

  Sheet 1 "Unlabeled"  — active entities with NO side assigned. Reviewer fills
                          the ASSIGNED_SIDE column.
  Sheet 2 "Combined"   — still-active multi-party (un-split) entities, with an
                          IS_ORPHAN flag (0 refs = safe to delete).

Context columns (chunks / edges referencing the entity) let the reviewer focus
on the ones that actually appear in the data. Read-only on the DB.

  python -m scripts.export_entity_review
Output: entity_review_for_ceo.xlsx in the project root.
"""
from __future__ import annotations

import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config.settings import Settings
from src.db.mongo import MongoClientWrapper
from src.utils.logger import logger

SIDE_OPTIONS = "david_network | our_side | co_victim | third_party | unknown"

_NUMBER_ONLY = re.compile(r"^\s*\d+\s+(?:llc|inc|corp|l\.?l\.?c\.?)\b", re.I)
_ADDR_CODED = re.compile(r"^\s*\d+\s*[A-Za-z]")


def _is_combined(name: str) -> bool:
    up = (name or "").upper()
    return (" & " in up) or (" AND " in up) or (up.count(",") >= 2)


def _why_unlabeled(name: str, refs_total: int) -> str:
    """Plain-language reason this entity has no side yet — so the reviewer
    knows what to do with it."""
    up = (name or "").strip()
    low = up.lower()
    if _is_combined(up):
        return ("Combined multi-party record (several names in one) — must be "
                "SPLIT into each party first; each part then gets its own side.")
    if refs_total == 0:
        return ("Orphan: no documents or relationships reference it — likely an "
                "OCR/parse artifact. Safe to delete.")
    if any(h in low for h in ("mango tree", "mangotree")):
        return "Mango Tree entity — OURS (our_side)."
    if any(h in low for h in ("ipa", "island propert", "directional lending")):
        return "David network (IPA / Island Properties / Directional) — david_network."
    if _NUMBER_ONLY.match(up):
        return ("Number-only company name (e.g. '2034 LLC') — ambiguous: could be "
                "a David address-coded LLC or a third-party. Confirm.")
    if _ADDR_CODED.match(up):
        return ("Address-coded but an Inc/Corp (not a clear LLC) — confirm whether "
                "it is David's shell or a genuine third-party business.")
    return ("Neutral counterparty (prior owner / seller / HOA / bank / individual) "
            "— never classified. Default third_party unless it's actually ours, "
            "David's, or Brian's.")


def main() -> int:
    s = Settings.load()
    m = MongoClientWrapper(s.mongo_uri, s.mongo_db_name)
    ents = m.db["entities"]
    rels = m.db["relationships"]
    chunks = m.db["email_chunks_v2"]

    all_ents = {e["_id"]: e for e in ents.find(
        {}, {"canonical_name": 1, "canonical_address": 1, "kind": 1, "side": 1,
             "is_david": 1, "aliases": 1, "is_active": 1})}
    active = [e for e in all_ents.values() if e.get("is_active") is not False]

    # property id -> human address (for context)
    def label(eid):
        e = all_ents.get(eid, {})
        return e.get("canonical_address") or e.get("canonical_name") or eid

    # Load all edges once; build entity -> linked PROPERTY addresses (so the
    # reviewer sees "this person/LLC is tied to 12 Mallard Path").
    linked_props: dict = {}
    for r in rels.find({}, {"src": 1, "dst": 1}):
        for a, b in ((r.get("src"), r.get("dst")), (r.get("dst"), r.get("src"))):
            if all_ents.get(b, {}).get("kind") == "property":
                linked_props.setdefault(a, set()).add(label(b))

    def context(eid):
        props = sorted(linked_props.get(eid, []))[:5]
        if props:
            return "; ".join(props)
        # no graph link -> pull a sample document this entity appears in
        ch = chunks.find_one({"entity_ids": eid},
                             {"property_address": 1, "subject": 1, "filename": 1})
        if ch:
            return (ch.get("property_address") or ch.get("subject")
                    or ch.get("filename") or "")[:80]
        return "(no documents reference this — orphan)"

    def refs(eid):
        c = chunks.count_documents({"entity_ids": eid})
        e = rels.count_documents({"$or": [{"src": eid}, {"dst": eid}]})
        return c, e

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    fill_in = PatternFill("solid", fgColor="FFF2CC")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill

    # Only PARTIES (person/llc/org) need a side decision — properties take
    # their side from their owner, so they're excluded entirely.
    PARTY_KINDS = ("person", "llc", "org")

    # ---- Sheet 1: Parties to label (no side) ----
    ws1 = wb.active
    ws1.title = "Parties to label"
    cols1 = ["name", "kind", "linked_property / context", "WHY unlabeled",
             "aliases", "chunks_referencing", "edges",
             f"ASSIGNED_SIDE ({SIDE_OPTIONS})", "notes", "entity_id"]
    ws1.append(cols1)
    style_header(ws1, len(cols1))
    # show the most-referenced (most important) first
    no_side = [e for e in active if not e.get("side")
               and e.get("kind") in PARTY_KINDS]
    no_side_ref = [(e, refs(e["_id"])) for e in no_side]
    no_side_ref.sort(key=lambda t: (t[1][0] + t[1][1]), reverse=True)
    n1 = 0
    for e, (c, ed) in no_side_ref:
        ws1.append([e.get("canonical_name") or e.get("canonical_address"),
                    e.get("kind"), context(e["_id"]),
                    _why_unlabeled(e.get("canonical_name") or "", c + ed),
                    ", ".join((e.get("aliases") or [])[:6]),
                    c, ed, "", "", e["_id"]])
        n1 += 1
    for r in range(2, n1 + 2):  # highlight the fill-in column
        ws1.cell(row=r, column=8).fill = fill_in
    ws1.freeze_panes = "A2"

    # ---- Sheet 2: Combined (un-split, still active) ----
    ws2 = wb.create_sheet("Combined")
    cols2 = ["combined_name", "kind", "current_side", "linked_property / context",
             "WHY (needs split)", "chunks_referencing", "edges",
             "IS_ORPHAN(0 refs = safe to delete)",
             "SPLIT_INTO / who-belongs-to-whom", "notes", "entity_id"]
    ws2.append(cols2)
    style_header(ws2, len(cols2))
    combined = [e for e in active if _is_combined(e.get("canonical_name") or "")
                and e.get("kind") in PARTY_KINDS]
    combined_ref = [(e, refs(e["_id"])) for e in combined]
    combined_ref.sort(key=lambda t: (t[1][0] + t[1][1]), reverse=True)
    n2 = 0
    for e, (c, ed) in combined_ref:
        why = ("Orphan combined record (0 refs) — safe to delete." if (c + ed) == 0
               else "Several parties bundled in one record — split into each "
                    "named party, then assign each its own side.")
        ws2.append([e.get("canonical_name"), e.get("kind"), e.get("side"),
                    context(e["_id"]), why, c, ed,
                    "YES" if (c + ed) == 0 else "no", "", "", e["_id"]])
        n2 += 1
    for r in range(2, n2 + 2):
        ws2.cell(row=r, column=9).fill = fill_in
    ws2.freeze_panes = "A2"

    # widen columns a bit
    for ws in (ws1, ws2):
        for col in ws.columns:
            width = min(60, max(12, max(len(str(c.value or "")) for c in col) + 2))
            ws.column_dimensions[col[0].column_letter].width = width

    out = "entity_review_for_ceo_v4.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        import time
        out = f"entity_review_for_ceo_{int(time.time())}.xlsx"
        wb.save(out)
    logger.info(f"wrote {out}: Unlabeled={n1} rows, Combined={n2} rows")
    m.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
