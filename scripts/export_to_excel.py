"""
Export emails from MongoDB to Excel files (XLSX), 500 rows per file by default.

Output:
    exports/emails_part_01_of_NN_<from>-<to>.xlsx
    exports/emails_part_02_of_NN_<from>-<to>.xlsx
    ...

Each row contains the full clean dataset for one email (subject, parties,
dates, body, folder, attachment list, etc.).

Usage:
    python scripts/export_to_excel.py
    python scripts/export_to_excel.py --rows-per-file 500
    python scripts/export_to_excel.py --order asc      # oldest first (default)
    python scripts/export_to_excel.py --order desc     # newest first
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Iterable

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import Settings
from src.db.mongo import MongoClientWrapper
from src.utils.logger import configure_logger, logger

# ---------------------------------------------------------------------------
# Excel layout
# ---------------------------------------------------------------------------
COLUMNS = [
    ("row_num", 8),
    ("date", 20),
    ("date_sent", 20),
    ("date_received", 20),
    ("from_name", 28),
    ("from_email", 36),
    ("to", 50),
    ("cc", 50),
    ("bcc", 50),
    ("subject", 60),
    ("folder", 30),
    ("importance", 12),
    ("has_attachments", 14),
    ("attachment_count", 14),
    ("attachment_filenames", 60),
    ("body_format", 12),
    ("body_text", 120),                # cleaned body
    ("body_text_raw_preview", 80),     # first 1500 chars of raw, for verification
    ("body_chars", 12),
    ("thread_id", 30),
    ("internet_message_id", 40),
    ("pst_entry_id", 14),
    ("mongo_id", 26),
]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=12, color="1F4E78")

# Cells in Excel can hold at most 32,767 chars
EXCEL_CELL_MAX = 32_000


def _fmt_address_list(addrs: list | None) -> str:
    if not addrs:
        return ""
    parts = []
    for a in addrs:
        name = (a or {}).get("name", "")
        email = (a or {}).get("email", "")
        if name and email:
            parts.append(f"{name} <{email}>")
        elif email:
            parts.append(email)
        elif name:
            parts.append(name)
    return "; ".join(parts)


def _fmt_dt(dt) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return str(dt)


def _truncate(text: str | None, limit: int = EXCEL_CELL_MAX) -> str:
    if not text:
        return ""
    if len(text) <= limit:
        return text
    return text[: limit - 50] + "\n...[truncated for excel cell]"


def _safe_str(v) -> str:
    if v is None:
        return ""
    return str(v)


def _build_row(idx: int, doc: dict, attachments_by_email: dict) -> list:
    sender = doc.get("from") or {}
    att_filenames = ""
    att_docs = attachments_by_email.get(doc["_id"], [])
    if att_docs:
        att_filenames = "; ".join(a.get("filename", "") for a in att_docs)

    body = doc.get("body_text", "") or ""
    raw = doc.get("body_text_raw", "") or ""
    raw_preview = raw[:1500] + ("…" if len(raw) > 1500 else "")

    return [
        idx,
        _fmt_dt(doc.get("date")),
        _fmt_dt(doc.get("date_sent")),
        _fmt_dt(doc.get("date_received")),
        _safe_str(sender.get("name")),
        _safe_str(sender.get("email")),
        _fmt_address_list(doc.get("to")),
        _fmt_address_list(doc.get("cc")),
        _fmt_address_list(doc.get("bcc")),
        _safe_str(doc.get("subject")),
        _safe_str(doc.get("folder_path")),
        _safe_str(doc.get("importance")),
        bool(doc.get("has_attachments")),
        int(doc.get("attachment_count") or 0),
        att_filenames,
        _safe_str(doc.get("body_format")),
        _truncate(body),
        _truncate(raw_preview),
        len(body),
        _safe_str(doc.get("thread_id")),
        _safe_str(doc.get("internet_message_id")),
        _safe_str(doc.get("pst_entry_id")),
        _safe_str(doc.get("_id")),
    ]


_PROTECTION_TAG = re.compile(rb"<workbookProtection\s*/>")


def _post_process_xlsx(out_path: Path) -> None:
    """Make the .xlsx 100% editable in Excel by removing leftover lock-related
    XML that openpyxl emits even when nothing is protected.

    Specifically removes the empty `<workbookProtection/>` element that Excel
    otherwise interprets as "structure protected, no password" — which can
    cause Excel to silently disable Insert/Delete/Move-Sheet operations and
    leave the user thinking the file is read-only.
    """
    tmp = out_path.with_suffix(out_path.suffix + ".tmp")
    with zipfile.ZipFile(out_path, "r") as src, zipfile.ZipFile(
        tmp, "w", zipfile.ZIP_DEFLATED
    ) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/workbook.xml":
                data = _PROTECTION_TAG.sub(b"", data)
            dst.writestr(item, data)
    out_path.unlink()
    tmp.rename(out_path)


def _write_workbook(out_path: Path, rows: list[list], part_idx: int, total_parts: int, range_label: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Emails {part_idx} of {total_parts}"

    # Header row (no merged title row — merged cells confuse users into
    # thinking the workbook is locked).
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes so the header stays visible.  No sheet protection.
    ws.freeze_panes = "A2"
    ws.protection.sheet = False
    # Explicitly clear every protection toggle (openpyxl defaults some to True
    # even when sheet=False — Excel can render that as a "soft lock").
    for attr in ("formatCells", "formatColumns", "formatRows",
                 "insertColumns", "insertRows", "insertHyperlinks",
                 "deleteColumns", "deleteRows", "selectLockedCells",
                 "sort", "autoFilter", "pivotTables", "selectUnlockedCells"):
        if hasattr(ws.protection, attr):
            setattr(ws.protection, attr, False)

    wrap = Alignment(wrap_text=True, vertical="top")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = wrap
        ws.row_dimensions[r_idx].height = 60

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # Post-process the saved file to strip the empty <workbookProtection/>
    # element openpyxl emits unconditionally.  This is the root cause of
    # Excel showing the file as "needs Enable Editing" / not letting users
    # type in cells.
    _post_process_xlsx(out_path)


def export(settings: Settings, mongo: MongoClientWrapper, rows_per_file: int, order: str) -> None:
    settings.export_dir.mkdir(parents=True, exist_ok=True)
    sort_dir = 1 if order == "asc" else -1

    total = mongo.emails.count_documents({})
    if total == 0:
        logger.error("No emails in database to export.")
        return

    total_parts = (total + rows_per_file - 1) // rows_per_file
    logger.info(
        f"Exporting {total:,} emails to {total_parts} Excel file(s) "
        f"of up to {rows_per_file} rows each "
        f"(order: {order}) into {settings.export_dir}"
    )

    # 1) Pre-load attachment filenames grouped by email — ONE query, no N+1.
    logger.info("Pre-loading attachment filenames…")
    attachments_by_email: dict = {}
    for a in mongo.attachments.find({}, {"email_id": 1, "filename": 1}):
        eid = a.get("email_id")
        if eid is None:
            continue
        attachments_by_email.setdefault(eid, []).append(
            {"filename": a.get("filename", "")}
        )
    logger.info(f"Loaded attachments for {len(attachments_by_email):,} emails.")

    # 2) Stream emails sorted by indexed `date` field (no in-memory sort).
    logger.info("Streaming emails…")
    cursor = mongo.emails.find(
        {},
        batch_size=200,
    ).sort([("date", sort_dir)]).hint("ix_date")

    rows: list[list] = []
    part_idx = 1
    global_idx = 0
    first_date_in_part = None
    last_date_in_part = None

    try:
        for doc in cursor:
            global_idx += 1
            d = doc.get("date") or doc.get("date_sent")
            if first_date_in_part is None:
                first_date_in_part = d
            last_date_in_part = d

            rows.append(_build_row(global_idx, doc, attachments_by_email))

            if len(rows) >= rows_per_file:
                _flush_part(settings, rows, part_idx, total_parts,
                            first_date_in_part, last_date_in_part)
                part_idx += 1
                rows = []
                first_date_in_part = None
                last_date_in_part = None

        if rows:
            _flush_part(settings, rows, part_idx, total_parts,
                        first_date_in_part, last_date_in_part)
    finally:
        cursor.close()

    logger.info(f"Done. {global_idx:,} rows written across {total_parts} file(s).")


def _flush_part(settings: Settings, rows: list[list], part_idx: int, total_parts: int,
                first_date, last_date) -> None:
    def _label(d):
        if isinstance(d, datetime):
            return d.strftime("%Y-%m-%d")
        return "n-a"

    range_label = f"{_label(first_date)}_to_{_label(last_date)}"
    fname = f"emails_part_{part_idx:02d}_of_{total_parts:02d}_{range_label}.xlsx"
    out = settings.export_dir / fname
    _write_workbook(out, rows, part_idx, total_parts, range_label.replace("_", " "))
    logger.info(f"  wrote {out.name} ({len(rows)} rows)")


def main() -> int:
    parser = argparse.ArgumentParser(description="Export emails to Excel.")
    parser.add_argument("--rows-per-file", type=int, default=None, help="Override EXCEL_ROWS_PER_FILE")
    parser.add_argument("--order", choices=("asc", "desc"), default="asc",
                        help="Sort order by date (default: asc / oldest first)")
    args = parser.parse_args()

    settings = Settings.load()
    configure_logger(settings.logs_dir)

    rows_per_file = args.rows_per_file or settings.excel_rows_per_file
    if rows_per_file <= 0:
        logger.error("rows_per_file must be > 0")
        return 2

    mongo = MongoClientWrapper(settings.mongo_uri, settings.mongo_db_name)
    try:
        mongo.ping()
        export(settings, mongo, rows_per_file, args.order)
        return 0
    finally:
        mongo.close()


if __name__ == "__main__":
    raise SystemExit(main())
