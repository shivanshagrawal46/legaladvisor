"""Correct the DOS filing date of 11 H LLC: 2029-03-14 (typo) -> 2019-03-14.

A formation date in the FUTURE (2029) makes the anachronism detector think the
LLC took title before it existed — a FALSE 'backdating' finding. Fixing the
year removes that false positive. Updates both the source Excel and the live
entity, then re-runs the detectors so findings refresh.

  python -m scripts.fix_11h_dos_date            # DRY-RUN (show current values)
  python -m scripts.fix_11h_dos_date --live     # apply fix + re-run detectors
"""
from __future__ import annotations

import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from config.settings import Settings
from src.db.mongo import MongoClientWrapper
from src.utils.logger import logger

XLSX = Path(r"C:\Users\SHIVANSH AGRAWAL\Desktop\outlook_attachments\List of LLC formed.xlsx")
WRONG_YEAR, RIGHT_YEAR = 2029, 2019
NAME_RE = re.compile(r"^\s*11\s*h\b", re.I)   # "11 H LLC" / "11H LLC"


def main() -> int:
    live = "--live" in sys.argv
    now = datetime.now(timezone.utc)
    s = Settings.load()
    m = MongoClientWrapper(s.mongo_uri, s.mongo_db_name)
    ents = m.db["entities"]

    # --- DB entity ---
    ent = ents.find_one({"kind": "llc", "canonical_name": {"$regex": r"^\s*11\s*H", "$options": "i"}})
    if ent:
        cur = ent.get("dos_filing_date")
        logger.info(f"DB entity {ent['_id']} ({ent.get('canonical_name')!r}) "
                    f"dos_filing_date = {cur}")
        if isinstance(cur, datetime) and cur.year == WRONG_YEAR:
            fixed = cur.replace(year=RIGHT_YEAR)
            logger.info(f"  -> will correct to {fixed}")
            if live:
                ents.update_one({"_id": ent["_id"]}, {"$set": {
                    "dos_filing_date": fixed, "dos_date_corrected": True,
                    "updated_at": now}})
        else:
            logger.info("  (DB value is not 2029 — nothing to change there)")
    else:
        logger.warning("11 H LLC entity not found in DB")

    # --- source Excel ---
    if XLSX.exists():
        wb = load_workbook(str(XLSX))
        ws = wb.active
        changed = False
        for row in ws.iter_rows(min_row=1):
            c0 = row[0].value
            if c0 and NAME_RE.match(str(c0)) and len(row) > 1:
                cell = row[1]
                v = cell.value
                logger.info(f"Excel row {row[0].row}: owner={c0!r} filing_cell={v!r}")
                if isinstance(v, datetime) and v.year == WRONG_YEAR:
                    cell.value = v.replace(year=RIGHT_YEAR)
                    changed = True
                    logger.info(f"  -> corrected Excel cell to {cell.value}")
                elif isinstance(v, str) and "2029" in v:
                    cell.value = v.replace("2029", "2019")
                    changed = True
                    logger.info(f"  -> corrected Excel cell to {cell.value!r}")
        if live and changed:
            try:
                wb.save(str(XLSX))
                logger.info("  Excel saved.")
            except PermissionError:
                logger.warning("  Excel is OPEN — close it and re-run to save the "
                               "source file. (DB already corrected regardless.)")
    else:
        logger.warning(f"Excel not found at {XLSX}")

    if live:
        logger.info("Re-running detectors to refresh findings...")
        from scripts.run_detectors import main as run_detectors_main
        try:
            run_detectors_main()
        except SystemExit:
            pass
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"detector re-run note: {exc}")
    else:
        logger.info("DRY-RUN — re-run with --live to apply + refresh findings.")
    m.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
