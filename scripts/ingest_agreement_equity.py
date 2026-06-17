"""
Ingest two David-matter documents:
  1) MangoTree Service Agreement with Island Property (PDF, scanned) — the
     operative contract between US (Mango Tree) and David's Island Properties.
     Full Claude Vision OCR (-> GPT-5 -> RapidOCR fallback), stored as a
     documents/ row (source_type=service_agreement), parties linked.
  2) Equity-in-properties spreadsheet (xlsx) — David's equity per property.
     Parsed cell-by-cell (NOT OCR — it is structured data), stored as a
     documents/ row (source_type=equity_schedule) with structured rows, and
     each row's equity/mortgage/lis-pendens/fraud facts pushed onto the
     matching property entity + linked to David.

Usage:  python -m scripts.ingest_agreement_equity --live   (default dry-run)
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

import config.settings  # noqa: F401
from config.settings import Settings
from src.db.mongo import MongoClientWrapper
from src.rag.evidence_schema import DEFAULT_MATTER_ID
from src.utils.hashing import sha256_bytes
from src.utils.logger import logger
from scripts.build_entities_from_llc import norm_name, norm_addr
from scripts.ingest_title_reports import resolve_owner_entity

ROOT = Path(r"C:\Users\SHIVANSH AGRAWAL\Desktop\outlook_attachments")
AGREEMENT_PDF = ROOT / "MangoTree Service Agreement with Island Property.pdf"
EQUITY_XLSX = ROOT / "IPA properties for sheriff sale - revised - Yvonne Mar 25, 2025.xlsx"


def _num(v: Any) -> Optional[float]:
    if v is None or str(v).strip() in ("", "-"):
        return None
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(v)))
    except ValueError:
        return None


def parcel_digits(p: str) -> str:
    return re.sub(r"\D", "", p or "")


def ingest_agreement(docs, ents, s: Settings, now, dry: bool):
    from scripts.ingest_titles_full import full_ocr, strip_watermarks
    logger.info(f"OCR (Claude Vision) service agreement: {AGREEMENT_PDF.name}")
    if dry:
        logger.info("  [dry-run] would OCR + store")
        return
    res = full_ocr(AGREEMENT_PDF, s, force=True)
    pages = res.pages or []
    text = "\n\n".join(strip_watermarks(p.text or "") for p in pages if p.text).strip()
    methods: Dict[str, int] = {}
    for p in pages:
        methods[p.method] = methods.get(p.method, 0) + 1
    data = AGREEMENT_PDF.read_bytes()

    # link parties: David's Island Properties + our Mango Tree
    david = resolve_owner_entity(ents, "Island Properties & Associates LLC", now)
    mango_id = "ent_org_mangotree"
    ents.update_one({"_id": mango_id}, {"$set": {
        "_id": mango_id, "kind": "org", "matter_id": DEFAULT_MATTER_ID,
        "canonical_name": "Mango Tree (Rakesh Sir's team)", "name_norm": norm_name("Mango Tree"),
        "aliases": ["Mango Tree", "MangoTree Real Estate Holdings LP"],
        "is_david": False, "is_ours": True, "source": "service_agreement", "updated_at": now,
    }, "$setOnInsert": {"created_at": now}}, upsert=True)

    doc = {
        "_id": "doc_agreement_mangotree_island",
        "source_type": "service_agreement", "instrument_subtype": "service_agreement",
        "matter_id": DEFAULT_MATTER_ID,
        "corpus": "contract_records", "privilege_status": "non_privileged",
        "evidentiary_class": "operative_agreement", "authority_score": 1.08,
        "is_signed": ("signature" in text.lower() or "agreed" in text.lower()),
        "parties": [
            {"role": "service_provider_or_party", "name_raw": "Mango Tree", "entity_id": mango_id},
            {"role": "counterparty", "name_raw": "Island Properties & Associates",
             "entity_id": david["entity_id"]},
        ],
        "counterparty_is_david": True,
        "page_count": len(pages), "extraction_method": methods,
        "ocr_confidence": res.avg_ocr_confidence,
        "extracted_text": text,
        "custody": {"source_files": [AGREEMENT_PDF.name], "sha256": sha256_bytes(data),
                    "origin": "service_agreement", "ingested_at": now},
        "quality": {"needs_review": len(text) < 500},
        "updated_at": now, "created_at": now,
    }
    docs.update_one({"_id": doc["_id"]}, {"$set": doc}, upsert=True)
    logger.info(f"  stored service_agreement: pages={len(pages)} chars={len(text)} methods={methods}")


def _money_str(v) -> str:
    return f"${v:,.2f}" if isinstance(v, (int, float)) else "n/a"


def _equity_row_block(x: Dict[str, Any]) -> str:
    """Table-aware, self-contained record for ONE property row.

    Address-led (so a per-property query retrieves it) and fully LABELLED
    (so column meaning + amounts are unambiguous to retrieval, the verifier,
    and the reader) — instead of a flat 'k=v k=v' line that loses table
    structure when chunked."""
    addr = ", ".join(p for p in [x.get("street"), x.get("city"),
                                 x.get("state")] if p) or "(address n/a)"
    lines = [
        f"### Equity schedule — Property #{x.get('serial')}: {addr}",
        f"Record owner: {x.get('owner_name') or 'n/a'}",
        f"Parcel / APN: {x.get('parcel') or 'n/a'}  |  County: {x.get('county') or 'n/a'}",
        f"Market value (John est.): {_money_str(x.get('mkt_value_john'))}  |  "
        f"Market value (Zillow): {_money_str(x.get('mkt_value_zillow'))}",
        f"Mortgage balance: {_money_str(x.get('mortgage'))}  |  Lender: {x.get('lender') or 'n/a'}",
        f"Real-estate taxes owed: {_money_str(x.get('re_taxes_owed'))}  |  "
        f"Property tax: {_money_str(x.get('property_tax'))}",
        f"Equity: {_money_str(x.get('equity'))}",
        f"Lis pendens: {x.get('lis_pendens') or 'none noted'}  |  "
        f"Active foreclosure: {x.get('active_foreclosure') or 'none noted'}",
        f"Judgment: {x.get('judgement') or 'none noted'}  |  "
        f"Fraudulent flag: {x.get('fraudulent') or 'no'}",
    ]
    if x.get("comments_2025"):
        lines.append(f"Comments (2025): {x['comments_2025']}")
    return "\n".join(lines)


def ingest_equity(docs, ents, s: Settings, now, dry: bool):
    wb = load_workbook(str(EQUITY_XLSX), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [("" if c is None else str(c)).strip() for c in rows[0]]
    logger.info(f"Equity Excel: {len(rows)} rows; columns: {header[:16]}")

    equity_rows: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if not r:
            continue
        serial = ("" if r[0] is None else str(r[0])).strip()
        owner = ("" if len(r) < 2 or r[1] is None else str(r[1])).strip()
        if not re.match(r"^\d+$", serial) or not owner:
            continue  # skip legend / blank / category rows
        rec = {
            "serial": serial, "owner_name": owner,
            "street": str(r[2]).strip() if len(r) > 2 and r[2] else None,
            "city": str(r[3]).strip() if len(r) > 3 and r[3] else None,
            "state": str(r[4]).strip() if len(r) > 4 and r[4] else None,
            "county": str(r[5]).strip() if len(r) > 5 and r[5] else None,
            "parcel": str(r[6]).strip() if len(r) > 6 and r[6] else None,
            "title_report_link": str(r[7]).strip() if len(r) > 7 and r[7] else None,
            "mkt_value_john": _num(r[8]) if len(r) > 8 else None,
            "mkt_value_zillow": _num(r[9]) if len(r) > 9 else None,
            "property_tax": _num(r[10]) if len(r) > 10 else None,
            "lis_pendens": str(r[11]).strip() if len(r) > 11 and r[11] else None,
            "mortgage": _num(r[12]) if len(r) > 12 else None,
            "re_taxes_owed": _num(r[13]) if len(r) > 13 else None,
            "equity": _num(r[14]) if len(r) > 14 else None,
            "lender": str(r[15]).strip() if len(r) > 15 and r[15] else None,
            "comments_2025": str(r[24]).strip() if len(r) > 24 and r[24] else None,
            "active_foreclosure": str(r[25]).strip() if len(r) > 25 and r[25] else None,
            "judgement": str(r[27]).strip() if len(r) > 27 and r[27] else None,
            "fraudulent": str(r[28]).strip() if len(r) > 28 and r[28] else None,
        }
        equity_rows.append(rec)
    logger.info(f"  parsed {len(equity_rows)} property equity rows")
    tot_equity = sum(x["equity"] for x in equity_rows if x.get("equity"))
    logger.info(f"  total equity across rows: ${tot_equity:,.0f}")

    if dry:
        for x in equity_rows[:5]:
            logger.info(f"   {x['serial']}: {x['owner_name']} | {x['street']} | parcel={x['parcel']} | "
                        f"equity={x['equity']} mortgage={x['mortgage']} fraud={x['fraudulent']}")
        return

    data = EQUITY_XLSX.read_bytes()
    linked = 0
    for x in equity_rows:
        # link to property entity by parcel digits, else by normalized address
        pid = None
        if x.get("parcel"):
            pdg = parcel_digits(x["parcel"])
            hit = ents.find_one({"kind": "property", "parcel_digits_idx": pdg}, {"_id": 1}) \
                if pdg else None
            if not hit and pdg:
                for e in ents.find({"kind": "property"}, {"_id": 1, "parcel_id": 1}):
                    if parcel_digits(e.get("parcel_id") or "") == pdg:
                        hit = e
                        break
            if hit:
                pid = hit["_id"]
        if not pid and x.get("street"):
            an = norm_addr(f"{x['street']} {x.get('city') or ''}")
            hit = ents.find_one({"kind": "property", "address_norm": an}, {"_id": 1})
            if hit:
                pid = hit["_id"]
        x["property_entity_id"] = pid
        owner_res = resolve_owner_entity(ents, x["owner_name"], now)
        x["owner_entity_id"] = owner_res["entity_id"]
        if pid:
            linked += 1
            ents.update_one({"_id": pid}, {"$set": {
                "equity": x.get("equity"), "mkt_value": x.get("mkt_value_john") or x.get("mkt_value_zillow"),
                "mortgage_amount": x.get("mortgage"), "re_taxes_owed": x.get("re_taxes_owed"),
                "lis_pendens": x.get("lis_pendens"), "lender": x.get("lender"),
                "active_foreclosure": x.get("active_foreclosure"),
                "fraudulent_flag": x.get("fraudulent"),
                "equity_source": EQUITY_XLSX.name, "equity_as_of": "2025-03-25",
                "updated_at": now,
            }})

    doc = {
        "_id": "doc_equity_schedule_2025_03_25",
        "source_type": "equity_schedule", "instrument_subtype": "equity_schedule",
        "matter_id": DEFAULT_MATTER_ID,
        "corpus": "financial_records", "privilege_status": "work_product",
        "evidentiary_class": "internal_analysis", "authority_score": 1.06,
        "title": "IPA properties for sheriff sale (David equity) — Yvonne, Mar 25 2025",
        "as_of_date": datetime(2025, 3, 25, tzinfo=timezone.utc),
        "row_count": len(equity_rows), "equity_rows": equity_rows,
        "total_equity": tot_equity,
        # Table-aware text: one labelled, address-led block per property row so
        # chunking aligns to row boundaries and column meaning survives.
        "extracted_text": "\n\n".join(_equity_row_block(x) for x in equity_rows),
        "custody": {"source_files": [EQUITY_XLSX.name], "sha256": sha256_bytes(data),
                    "origin": "equity_schedule", "ingested_at": now},
        "quality": {"needs_review": False},
        "updated_at": now, "created_at": now,
    }
    docs.update_one({"_id": doc["_id"]}, {"$set": doc}, upsert=True)
    logger.info(f"  stored equity_schedule: {len(equity_rows)} rows, linked {linked} to property entities")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", default=True)
    ap.add_argument("--live", dest="dry_run", action="store_false")
    ap.add_argument("--equity-only", action="store_true",
                    help="re-ingest only the equity schedule (skip the costly "
                         "Claude-Vision re-OCR of the service agreement)")
    args = ap.parse_args()
    s = Settings.load()
    now = datetime.now(timezone.utc)
    if not args.dry_run and not args.equity_only:
        from src.extractor.claude_ocr import init_spend_guard
        init_spend_guard(s.ocr_vision_budget_usd)
    m = MongoClientWrapper(s.mongo_uri, s.mongo_db_name)
    m.ping()
    docs, ents = m.db["documents"], m.db["entities"]
    ingest_equity(docs, ents, s, now, args.dry_run)
    if not args.equity_only:
        ingest_agreement(docs, ents, s, now, args.dry_run)
    if args.dry_run:
        logger.info("DRY RUN — re-run with --live to store.")
    m.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
