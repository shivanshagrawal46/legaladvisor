"""List every property and the title reports we have ingested for it.

Produces:
  • a console summary
  • an Excel workbook (one row per stored title-report document) showing the
    property, the report type/date/vendor, version lineage, and the ORIGINAL
    source PDF file name(s) folded into that document (custody.source_files).

The source-file column is the important one for the user: where MORE THAN ONE
source file is folded into a single document, that is a place the old dedup may
have collapsed genuinely distinct dated reports — so the user can cross-check
against the folder and hand back any report that is missing.

Usage:  python -m scripts.list_ingested_titles
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config.settings import Settings
from src.db.mongo import MongoClientWrapper

OUT_XLSX = Path(__file__).resolve().parent.parent / "ingested_title_reports.xlsx"


def _d(v):
    try:
        return v.strftime("%Y-%m-%d")
    except Exception:
        return ""


def _key_date(doc) -> str:
    """The date that identifies this report: search/new-effective for an update,
    completed for a full search."""
    if doc.get("instrument_subtype") == "update_search" or doc.get("is_update"):
        return _d(doc.get("search_date") or doc.get("new_effective_date") or doc.get("effective_date"))
    return _d(doc.get("completed_date") or doc.get("effective_date"))


def _src_kind(path: str) -> str:
    """Classify a source filename as a full search or an update search."""
    low = (path or "").lower()
    if "update" in low or re.search(r"[-_ ]new\b", low):
        return "update"
    return "full"


def _collapse_flag(srcs) -> str:
    """Flag a stored document whose folded source files look like DIFFERENT
    reports were merged into one record (the dedup-collapse bug)."""
    if len(srcs) <= 1:
        return ""
    kinds = [_src_kind(s) for s in srcs]
    n_full, n_upd = kinds.count("full"), kinds.count("update")
    if n_full >= 1 and n_upd >= 1:
        return "LIKELY COLLAPSE (full+update merged)"
    if n_upd >= 2:
        return "REVIEW (multiple updates merged)"
    # multiple full-search files are usually the SAME report re-sent across
    # folders (2021 + 2026 'Sent to Dhibin'), so not flagged by default.
    return ""


def main() -> int:
    s = Settings.load()
    m = MongoClientWrapper(s.mongo_uri, s.mongo_db_name)
    m.ping()
    docs = m.db["documents"]

    rows = list(docs.find({"source_type": "title_report"}, {
        "property_address": 1, "address_norm": 1, "parcel_id": 1,
        "vendor": 1, "instrument_subtype": 1, "is_update": 1, "order_type": 1,
        "order_number": 1, "completed_date": 1, "index_date": 1, "search_date": 1,
        "effective_date": 1, "new_effective_date": 1, "old_effective_date": 1,
        "is_latest": 1, "version_group": 1, "version_count": 1, "version_index": 1,
        "custody": 1, "page_count": 1,
    }))

    # group by property (parcel preferred, else normalized address)
    by_prop = defaultdict(list)
    for d in rows:
        pkey = d.get("parcel_id") or d.get("address_norm") or d.get("property_address") or d["_id"]
        by_prop[pkey].append(d)

    # ---- console summary ----
    print(f"\nTotal stored title-report documents : {len(rows)}")
    print(f"Distinct properties with a title report: {len(by_prop)}\n")

    multi_src = 0
    flagged = []  # (addr, flag, srcs)
    table = []  # for excel
    for pkey, group in sorted(by_prop.items(), key=lambda kv: (kv[1][0].get("property_address") or "")):
        addr = group[0].get("property_address") or "(no address)"
        parcel = group[0].get("parcel_id") or ""
        group.sort(key=lambda d: (_key_date(d) or "9999"))
        print("=" * 78)
        print(f"{addr}   [parcel {parcel}]   — {len(group)} stored report(s)")
        for d in group:
            typ = "UPDATE" if (d.get("instrument_subtype") == "update_search" or d.get("is_update")) else "FULL"
            srcs = (d.get("custody") or {}).get("source_files") or []
            if len(srcs) > 1:
                multi_src += 1
            flag = _collapse_flag(srcs)
            if flag:
                flagged.append((addr, flag, srcs))
            ver = f"v{d.get('version_index','?')}/{d.get('version_count','?')}"
            latest = "  *LATEST*" if d.get("is_latest") else ""
            fmark = f"   <<< {flag}" if flag else ""
            if d.get("vendor") == "protitle":
                dd = (f"order#={d.get('order_number') or '-'} "
                      f"completed={_d(d.get('completed_date')) or '-'} "
                      f"index={_d(d.get('index_date')) or '-'}")
            else:
                dd = (f"order_type={d.get('order_type') or '-'} "
                      f"search={_d(d.get('search_date')) or '-'} "
                      f"old_eff={_d(d.get('old_effective_date')) or '-'} "
                      f"new_eff={_d(d.get('new_effective_date')) or '-'}")
            print(f"   - [{typ}] {d.get('vendor')} | {dd} | {ver}{latest}{fmark}")
            for sf in srcs:
                print(f"        source file: {sf}")
            table.append({
                "Property Address": addr,
                "Parcel": parcel,
                "Address Norm": d.get("address_norm") or "",
                "Report Type": typ,
                "Vendor": d.get("vendor"),
                # ProTitle dedup fields
                "Order #": d.get("order_number") or "",
                "Completed Date (PT)": _d(d.get("completed_date")),
                "Index Date (PT)": _d(d.get("index_date")),
                # Prowess dedup fields
                "Order Type": d.get("order_type") or "",
                "Search Date (PW)": _d(d.get("search_date")),
                "Old Eff Date (PW)": _d(d.get("old_effective_date")),
                "New Eff Date (PW)": _d(d.get("new_effective_date")),
                "Key Date": _key_date(d),
                "Version": ver,
                "Is Latest": "yes" if d.get("is_latest") else "",
                "Pages": d.get("page_count") or "",
                "# Source Files": len(srcs),
                "Collapse Flag": flag,
                "Source File(s)": " ; ".join(srcs),
            })

    print("\n" + "=" * 78)
    print(f"Documents that folded >1 source file: {multi_src}")
    print(f">>> LIKELY-COLLAPSE / REVIEW documents (distinct reports merged): {len(flagged)}")
    print("=" * 78)
    for addr, flag, srcs in flagged:
        print(f"  [{flag}] {addr}")
        for s in srcs:
            print(f"        {s}")

    # ---- Excel ----
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Title Reports"
        headers = list(table[0].keys()) if table else []
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="234A52")
            c.alignment = Alignment(vertical="center")
        for r in table:
            ws.append([r[h] for h in headers])
        # column widths
        widths = [34, 24, 30, 10, 9, 12, 16, 14, 14, 14, 14, 14, 12, 8, 9, 7, 12, 30, 70]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
        ws.freeze_panes = "A2"
        try:
            wb.save(OUT_XLSX)
            saved = OUT_XLSX
        except PermissionError:
            # target is open in Excel (file lock) — save to a fresh name instead
            from datetime import datetime as _dt
            saved = OUT_XLSX.with_name(f"ingested_title_reports_{_dt.now():%Y%m%d_%H%M%S}.xlsx")
            wb.save(saved)
            print(f"\n(NOTE: '{OUT_XLSX.name}' is open/locked — saved a fresh copy instead.)")
        print(f"\nExcel written: {saved}")
    except Exception as exc:  # noqa: BLE001
        print(f"\n(Excel not written: {exc}) — console list above is complete.")

    m.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
